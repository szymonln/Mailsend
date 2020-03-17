from django.shortcuts import render
import openpyxl

# Create your views here.
from django.http import HttpResponseRedirect
from django.shortcuts import render
from serialmail.models import MailProvider
from .forms import UploadFileForm, EmailProviderForm
import smtplib, ssl
from email.mime.text import MIMEText
import json


def root_view(request):
    return HttpResponseRedirect('/accounts/login')


def send_serial_mails(messages, config):
    port = 465  # For SSL

    # Create a secure SSL context
    context = ssl.create_default_context()

    with smtplib.SMTP_SSL(config['server'], port, context=context) as server:
        server.login(config['user'], config['password'])
        for msg in messages:
            message = MIMEText(msg["content"])
            message["Subject"] = msg["title"]
            message["From"] = config['user']
            message["To"] = msg["recipient"]
            server.sendmail(config['user'], msg["recipient"], message.as_string())


def handle_uploaded_file(f):
    with open(f.name, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)


def upload_file(request):
    if MailProvider.objects.filter(owner=request.user):
        provider = MailProvider.objects.filter(owner=request.user)[0]
        email_form = EmailProviderForm(instance=provider, initial={'password': provider.password})
    else:
        email_form = EmailProviderForm()
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if request.POST.get("user"):
            provider = MailProvider.objects.filter(owner=request.user)[0]
            provider.user = request.POST.get("user")
            provider.server = request.POST.get("server")
            provider.password = request.POST.get("password")
            provider.save()
            email_form = EmailProviderForm(instance=provider, initial={'password': provider.password})
            return render(request, 'upload.html', {'form': form, 'email_form': email_form, 'email_updated': True})

        if form.is_valid():
            handle_uploaded_file(request.FILES['file'])
            return HttpResponseRedirect('/send_mails/'+request.FILES['file'].name)
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form,
                                           'email_form': email_form})


def send_mails(request, file_name):
    excel_file = file_name

    # you may put validations here to check extension or file size

    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets
    worksheet = wb[wb.sheetnames[0]]

    excel_data = list()
    # iterating over the rows and
    # getting value from each cell in row
    keys = []
    for row in worksheet.iter_rows(min_row=2, max_row=2, min_col=3, values_only=True):
        for elem in row:
            if elem:
                keys.append(elem)

    emails = []
    for row in worksheet.iter_rows(min_row=3, max_col=2, values_only=True):
        for elem in row:
            if elem:
                emails.append(elem)

    users_data = []
    for row in worksheet.iter_rows(min_row=3, min_col=2, values_only=True):
        it = 0
        user_data = {}
        for cell in row:
            if cell:
                if "@" in str(cell):
                    user_data["email"] = cell
                    continue
                else:
                    user_data[keys[it]] = cell
                it += 1
        if user_data:
            users_data.append(user_data)

    for row in worksheet.iter_rows(values_only=True):
        row_data = list()
        for cell in row:
            row_data.append(str(cell))
        excel_data.append(row_data)

    if request.method == "POST" and request.POST.get("sendmail"):
        messages_json = json.loads(request.POST.get('messages').replace("\'", "\""))
        provider = MailProvider.objects.filter(owner=request.user)[0]
        config = {'password': provider.password, 'user': provider.user, 'server': provider.server}
        send_serial_mails(messages_json, config)
        return render(request, 'send_mails.html', {"excel_data": excel_data,
                                                   "keys": keys,
                                                   "emails": emails,
                                                   "emails_ready": True,
                                                   "emails_count": len(messages_json)})

    if request.method == 'POST':
        messages = []
        message_content = request.POST.get("mail_content")
        title = request.POST.get("title")

        for user in users_data:
            message = {}
            message["title"] = title
            content = message_content
            for key in keys:
                replace_string = '{'+key+'}'
                content = content.replace(replace_string, str(user.get(key)))
            message["content"] = content
            message["recipient"] = user["email"]
            messages.append(message)

        return render(request, 'send_mails.html', {"excel_data": excel_data,
                                                   "keys": keys,
                                                   "emails": emails,
                                                   "data_ready": True,
                                                   "messages": messages})


    return render(request, 'send_mails.html', {"excel_data": excel_data,
                                               "keys": keys,
                                               "emails": emails})
